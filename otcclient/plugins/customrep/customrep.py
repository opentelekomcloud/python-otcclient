#!/usr/bin/env python
# -*- coding: utf-8 -*-

# This file is part of OTC Tool released under MIT license.
# Copyright (C) 2016 T-systems Kurt Garloff, Zsolt Nagy

from otcclient.core.OtcConfig import OtcConfig 
from otcclient.utils import utils_http
from otcclient.utils import utils_http, utils_templates

from otcclient.core.otcpluginbase import otcpluginbase
from otcclient.core.pluginmanager import getplugin
import json
from datetime import timedelta, datetime

# otc customrep tenantusage --tenant "[{'projectid': 'xxxxxxxxxxxxx','rowdelta': '0'},{'projectid': 'yyyyyyyyyy','rowdelta': '0'}]" --file1="c:\1\report.xlsx" 

def getcoldatehour55(indate):
    col = 3  # initial value 
    col += (indate + timedelta(hours=-8)).hour
    return col

    
class customrep(otcpluginbase):
    ar = {}    
    flavors = None
        
    @staticmethod
    def otcOutputHandler(): 
        return getplugin(OtcConfig.OUTPUT_FORMAT)
 
    def otctype(self):
        return "func" 

    """    
    @staticmethod 
    def _get_quotas():
        # clouderver/limits quota         
        url = "https://" + OtcConfig.DEFAULT_HOST+ "/v1/" + OtcConfig.PROJECT_ID + "/cloudservers/limits"        
        ret = utils_http.get(url)
        quotas = json.loads(ret)    
        for q in quotas["absolute"]:
            pass     #totalRAMUsed   ,maxTotalCores  
        # quota 
        url = "https://" + OtcConfig.DEFAULT_HOST+ "/v1/" + OtcConfig.PROJECT_ID + "/quotas"
        ret = utils_http.get(url)
        quotas = json.loads(ret)
        for q in quotas["quotas"]:
            #print(q)
            pass
        #customrep.otcOutputHandler().print_output(ret, mainkey="quotas")    
    """                     

    @staticmethod 
    def _get_flavor(flavorid):
        if not customrep.flavors:   
            url = "https://" + OtcConfig.DEFAULT_HOST + "/v1/" + OtcConfig.PROJECT_ID + "/cloudservers/flavors"        
            ret = utils_http.get(url)
            customrep.flavors = json.loads(ret)            
        for f in customrep.flavors["flavors"]:
            if flavorid == f.get("id"):
                return  f        

    @staticmethod 
    def tenantusage():
        if not OtcConfig.TENANT:
            raise Exception("Tenant list mandatory\neg:[{'projectid': 'xxxxxxxxxxxxxxxxxxxx','rowdelta': '1'}]")
        if not OtcConfig.FILE1:
            raise Exception("Template file mandatory\neg: --file1 mytemplatefile1.xlsx")        
        
        tenants = json.loads(str(OtcConfig.TENANT).replace("'", "\""))
                             
        for tenant in tenants:        
            OtcConfig.PROJECT_ID = tenant["projectid"]
            sumvcpus = 0
            sumram = 0

            url = "https://" + OtcConfig.DEFAULT_HOST + "/v2/" + OtcConfig.PROJECT_ID + "/servers/detail"                        
            ret = utils_http.get(url)        
            servers = json.loads(ret)
            servercount = len(servers["servers"])
            
            for s in servers["servers"]:            
                f = customrep._get_flavor(s.get("flavor").get("id"))
                sumram += int(f.get("ram")) 
                sumvcpus += int(f.get("vcpus"))
    
            url = "https://" + OtcConfig.DEFAULT_HOST + "/v2/" + OtcConfig.PROJECT_ID + "/cloudvolumes" + "/detail"
            ret = utils_http.get(url)
            volumes = json.loads(ret)        
            volumesum = 0
            for v in volumes["volumes"]:
                if v["status"] == "in-use": 
                    volumesum += v.get("size")
                    
            print("SUMVMs:" , servercount) 
            print("SUM vCPU:", sumvcpus)
            print("SUM RAM:", sumram)     
            print("SUM Volume:", volumesum)

            datecolumn = getcoldatehour55(datetime.now())
            update_xlsx(
                        datecolumn,
                        int(tenant["rowdelta"]),
                        cpus=sumvcpus,
                        ram=sumram,
                        servercount=servercount,
                        volumesum=volumesum,
                        report_file=OtcConfig.FILE1
                        )        
                
def update_xlsx(col, rowdelta, cpus=None, ram=None, servercount=None, volumesum=None, report_file=None):
    from openpyxl import load_workbook     
    # Open an xlsx for writing
    wb = load_workbook(filename=report_file)
    
    ws = wb["TOTAL"]
    ws.cell(row=2 + rowdelta, column=col).value = cpus     
    ws.cell(row=20 + rowdelta, column=col).value = ram 
    ws.cell(row=11 + rowdelta, column=col).value = servercount     
    ws.cell(row=29 + rowdelta, column=col).value = volumesum

    # save the csb file
    wb.save(report_file)
                
                
